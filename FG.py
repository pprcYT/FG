import asyncio#임포트할 모듈 임포트

import discord

import openpyxl

import sys

import datetime

import re

from discord.ext import commands

import os, random, time

from discord.ext.commands import bot

idA, moneyA, timeA, give, ID, TIME = [], [], [], 0, 0, 0
owner = [562409981488660491]
bad = ["시발","문재앙","일베","대께문","조국","추미애","박근혜","이승만","이명박","노무현","새끼","느그애비","병신","좆까","느개비","ㅅㄲ","ㅆㄲ","ㅂㅅ","ㅄ","ㅅㅂ","ㅆㅂ","ㅆㅃ","시빨","시벌","씨발","씨벌","미띤","미친","야발","느금마","니애미","니애비","애미","애비","ㅗ","fuck","fuckyou","fuck you","C발","리발","좆","씹","보지","자지","보1지"]
client = discord.Client()


try: #만약 파일이 없으면 새로 만듦
    f = open("UserData.txt", "r")
except:
    f = open("UserData.txt", "w")
    f.close()
    f = open("UserData.txt", "r")
while True: #유저들 데이터를 읽음 데이터 형식 : 유저ID,가지고 있는 돈,돈받은 시간
    line = f.readline()
    if not line: break
    line = line.split(",")
    idA.append(line[0])
    moneyA.append(int(line[1]))
    timeA.append(int(line[2]))
f.close()

@client.event
async def on_ready():
    print(client.user.id)
    print("봇 로딩 완료")#봇 시작이라고 뜨게하기

game = discord.Game(".도움말 을 입력해보세요")#게임에서 python bot 하는중... 이라고 표

@client.event
async def on_message(message):
    t = time.ctime()
    msg = message.content
    id = message.author.id
    print("[" + str(t) + "] <" + str(id) + "> [" + str(msg) + "]")
    f = open("message.txt", 'a')
    data = "[" + str(t) + "] <" + str(id) + "> [" + str(msg) + "]"
    f.write(data)
    f.write("\n")
    f.close
    if message.content.startswith(".도움말"):
        embed = discord.Embed(title="FG도우미의 도움말 입니다.", color=0x00ff00)
        embed.add_field(name="FG 관련 도움말 :grey_question:", value="'.카톡' '.디스코드' 가 있습니다.", inline=True)
        embed.add_field(name="놀기 :speech_left:", value="배우기 관련: '.배워 <물어볼 말> <대답>' '.잊어 <물어봤던 말>'/도박 관련: '.도박 <코인>' '.올인' '.코인받기' ", inline=False)
        embed.add_field(name="서버장 명령어 :crown:", value="'.청소 <할 개수>")
        embed.add_field(name="기타:guitar:", value="'.핑'제작자: FG저스티스", inline=False)
        embed.set_footer(text="..더 발전할 예정")
        embed.set_thumbnail(url="https://encrypted-tbn0.gstatic.com/images?q=tbn%3AANd9GcQROgSkjjPNgbwVEZPylZ5vmXoPCh2Rw3823g&usqp=CAU")
        await message.channel.send(embed=embed)

    if message.content.startswith(".카톡"):
        embed =  discord.Embed(title="카카오톡 오픈채팅방 주소입니다.", color=0xFFFF33)
        embed.add_field(name="주소", value="https://open.kakao.com/o/gfzDimpc", inline=True)
        embed.set_thumbnail(url="https://search.pstatic.net/common/?src=http%3A%2F%2Fimgnews.naver.net%2Fimage%2F311%2F2013%2F12%2F02%2F1385986145171_59_20131202211502.jpg&type=sc960_832")
        await message.channel.send(embed=embed)

    if message.content.startswith(".디스코드"):
        embed =  discord.Embed(title="디스코드 초대 링크입니다.", color=0x8C8C8C)
        embed.add_field(name="주소", value="https://discord.gg/u8AVfe2", inline=True)
        embed.set_thumbnail(url="https://search.pstatic.net/common/?src=http%3A%2F%2Fblogfiles.naver.net%2FMjAyMDA4MTJfNjcg%2FMDAxNTk3MjMxMTA1ODUy.VNf1YaPNvNJ7Y9utFJdKMDGGBu0XJePIh90Zj59FKw0g.znlM0icJ8vp-275z7Wbg3lUw9nV8wfuD0XfBaIeZbFIg.PNG.desbey7%2Ficonfinder_91_Discord_logo_logos_4373196.png&type=sc960_832")
        await message.channel.send(embed=embed)
    if message.content.startswith('.핑'):
        embed = discord.Embed(description=f"", colour=discord.Colour(0x594841))
        embed.set_author(name=f"현재 핑은 {int((client.latency * 1000))}'ms 입니다.")
        await message.channel.send(embed=embed)

    msg = message.content
    if msg in bad:
        await message.delete()
        embed = discord.Embed(title="욕설 감지", description="익명 님의 욕설이 감지되었습니다.", color=0xff0000)
        embed.set_thumbnail(url="https://creazilla-store.fra1.digitaloceanspaces.com/emojis/52726/cross-mark-emoji-clipart-md.png")
        await message.channel.send(embed=embed)
############################################배우기############################################

    if message.content.startswith('.잊어'):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 1001):
            if sheet["A" + str(i)].value == str(memory[1]):
                sheet["A" + str(i)].value = "-"
                sheet["B" + str(i)].value = " "
                await message.channel.send("잊어버렸어요!")
                file.save('기억.xlsx')
                break

    if message.content.startswith(".배워"):
        file = openpyxl.load_workbook('기억.xlsx')
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 1001):
            if sheet["A"+str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("정상적으로 배웠어요!")
                await message.channel.send("현재 사용중인 데이터 저장용량 : " + str(i)+" / 1000")
                break
        file.save("기억.xlsx")


    if message.content.startswith(".FG"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 1001):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" + str(i)].value)
                break

    if message.content.startswith(".기억 초기화") or message.content.startswith(".기억초기화"):
        if message.author.id in owner:
            file = openpyxl.load_workbook("기억.xlsx")
            sheet = file.active
            for i in range(1, 1001):
                sheet["A"+str(i)].value = "-"
            embed = discord.Embed(title="FG 도우미봇 : 기억초기화", description=f"Bot Developer 권한으로 기억을 바꿨어요!\n \n 기억초기화가 정상적으로 완료되었어요!", timestamp=message.created_at,
            colour = discord.Colour.teal()    
        )
            await message.channel.send(embed=embed)
            file.save("기억.xlsx")
        else:
            embed = discord.Embed(title=f"FG도우미 봇 : Error", description=f"Bot Developer 등급보다 낮은 등급을 가지고 있습니다. \n \n {message.author.mention} 님의 등급 : User", timestamp=message.created_at,
            colour = discord.Colour.red()
    )
            embed.set_footer(text="개발자 등급이상만 사용할 수 있는 명령입니다.")        
            await message.channel.send(embed=embed)
            
##################################도박##################################
    if message.content == ".코인받기":
        ID = str(message.author.id)
        TIME = int(time.time())
        if ID in idA: #만약 등록된 ID라면
            if TIME - timeA[idA.index(ID)] < 30: #1시간이 안 지났을 때
                embed = discord.Embed(title='FG도우미 도박 : Error', description='30초 마다 받을 수 있습니다.', color=0xFF0000)
                await message.channel.send(embed=embed)
                raise ValueError #탈출
            elif TIME - timeA[idA.index(ID)] >= 30: #1시간이 지났을 때
                timeA[idA.index(ID)] = int(time.time())
        give = random.randrange(1,10)*random.randrange(1000,10000) # 줄 돈
        if ID in idA:
            moneyA[idA.index(ID)] += give
            f = open("UserData.txt", "w") #저장
            for i in range(0,len(idA),1):
                f.write(str(idA[i])+","+str(moneyA[i])+","+str(timeA[i])+"\n")
            f.close()
        elif not ID in idA:
            idA.append(ID)
            moneyA.append(give)
            timeA.append(int(time.time()))
            f = open("UserData.txt", "w") #저장
            for i in range(0,len(idA),1):
                f.write(str(idA[i])+","+str(moneyA[i])+","+str(timeA[i])+"\n")
            f.close()
        msg = str(give)+"코인을 받았습니다. 현재 FG코인 : "+str(moneyA[idA.index(ID)])+"코인"
        embed = discord.Embed(title='', description=msg, color=0x00FF00)
        await message.channel.send(embed=embed)
    if message.content.startswith(".도박"):
        ID = str(message.author.id)
        msg = message.content.split()
        if msg[1].isdecimal() == False: #만약 숫자가 아니라면
            embed = discord.Embed(title='FG도우미 도박 : Error', description='숫자외에 다른 내용이 들어가 있어요!', color=0xFF0000)
            await message.channel.send(embed=embed)
            raise ValueError
        msg[1] = int(msg[1])
        if not ID in idA or moneyA[idA.index(ID)] - int(msg[1]) < 0: #등록된 ID가 아니거나 돈이 부족하면
            embed = discord.Embed(title='FG도우미 도박 : Error', description='보유한 코인보다 높은코인을 거실수 없어요!', color=0xFF0000)
            await message.channel.send(embed=embed)
            raise ValueError #탈출
        moneyA[idA.index(ID)] -= msg[1]
        give = random.randrange(1,11)
        await asyncio.sleep(1)
        if give % 2 == 0:
            moneyA[idA.index(ID)] += give*msg[1]
            embed = discord.Embed(title="도박 성공!", description="도박을 성공하여 [ "+str(give)+"배 ] 의 코인을 얻었어요! \n \n 현재 FG코인 • "+str(moneyA[idA.index(ID)])+" 코인", timestamp=message.created_at,
            colour = discord.Colour.dark_green()       
    )
            embed.set_footer(text="도박 확률은 50 : 50이에요!")
            await message.channel.send(embed=embed)
        elif give % 2 != 0:
            embed = discord.Embed(title="도박 실패..", description="도박을 성공했으면 [ "+str(give)+"배 ] 의 코인을 얻을 수 있었는데 실패했네요.. \n \n 현재 FG코인 • "+str(moneyA[idA.index(ID)])+" 코인", timestamp=message.created_at,
            colour = discord.Colour.dark_red()       
    )
            embed.set_footer(text="도박 확률은 50 : 50이에요!")
            await message.channel.send(embed=embed)
        f = open("UserData.txt", "w") #저장
        for i in range(0,len(idA),1):
            f.write(str(idA[i])+","+str(moneyA[i])+","+str(timeA[i])+"\n")
        f.close()

    if message.content == ".코인":
        ID = str(message.author.id)
        if ID in idA: #만약 등록된 ID라면
            embed = discord.Embed(title=f'{message.author.name}님이 보유한 코인', description=str(moneyA[idA.index(ID)])+" 코인", color=0x118811)
            await message.channel.send(embed=embed)
        elif not ID in idA: #등록된 ID가 아니라면
            embed = discord.Embed(title='보유한 코인', description="0 원", color=0x118811)
            await message.channel.send(embed=embed)

    if message.content == ".올인":
        ID = str(message.author.id)
        if not ID in idA or moneyA[idA.index(ID)] <= 0: #만약 돈이 부족하면
            embed = discord.Embed(title='', description='코인이 부족합니다.', color=0xFF0000)
            await message.channel.send(embed=embed)
            raise ValueError
        give = random.randrange(2,10)
        await asyncio.sleep(1)
        if give % 2 == 0:
            moneyA[idA.index(ID)]*= give
            embed = discord.Embed(title="FG도우미 도박 올인 : 성공", description="올인을 성공하여 [ "+str(give)+"배 ] 의 코인을 얻었어요! \n \n 현재 FG코인 • "+str(moneyA[idA.index(ID)])+" 코인", timestamp=message.created_at,
            colour = discord.Colour.dark_green()       
    )
            embed.set_footer(text="올인 확률은 50 : 50이에요!")
            await message.channel.send(embed=embed)
        elif give % 2 != 0:
            moneyA[idA.index(ID)] = 0
            embed = discord.Embed(title="FG도우미 도박 올인 : 실패", description="올인을 성공했으면 [ "+str(give)+"배 ] 의 코인을 얻을 수 있었는데 실패했네요.. \n \n 현재 FG코인 • "+str(moneyA[idA.index(ID)])+" 코인", timestamp=message.created_at,
            colour = discord.Colour.dark_red()       
    )
            embed.set_footer(text="올인 확률은 50 : 50이에요!")
            await message.channel.send(embed=embed)
        f = open("UserData.txt", "w") #저장
        for i in range(0,len(idA),1):
            f.write(str(idA[i])+","+str(moneyA[i])+","+str(timeA[i])+"\n")
        f.close()
####################################청소####################################
    if message.content.startswith(".청소"):
        if message.author.guild_permissions.manage_messages or message.author.id in owner:
            try:
                amount = message.content[4:]
                await message.channel.purge(limit=int(amount)+1)
                embed = discord.Embed(title="청소 완료!", description=f"{message.author.mention}, **{amount}** 개의 메시지를 청소했어요.", timestamp=message.created_at,
                colour = discord.Colour.green()
    )
                embed.set_footer(text="FG도우미#2263", icon_url="https://search.pstatic.net/common/?src=http%3A%2F%2Fcafefiles.naver.net%2F20120402_196%2Fsarasa5621_1333335668105taVCH_PNG%2F%25C8%25DE%25C1%25F6%25C5%25EB.png&type=sc960_832")
                await message.channel.send(embed=embed)
            except ValueError:
                embed = discord.Embed(title="청소 실패!", description=f"{message.author.mention}, 청소는 다음값은 무조건 숫자여야 합니다.", timestamp=message.created_at, 
                colour=discord.Colour.orange()
    )
                embed.set_footer(text="FG도우미#2263", icon_url="https://search.pstatic.net/common/?src=http%3A%2F%2Fcafefiles.naver.net%2F20120402_196%2Fsarasa5621_1333335668105taVCH_PNG%2F%25C8%25DE%25C1%25F6%25C5%25EB.png&type=sc960_832")
                await message.channel.send(embed=embed)
        else:
                embed = discord.Embed(title="청소 실패!", description=f"{message.author.mention}, 청소를 실행할 권한이 없어요.", timestamp=message.created_at, 
                colour=discord.Colour.red()
    )
                embed.set_footer(text="FG도우미#2263", icon_url="https://search.pstatic.net/common/?src=http%3A%2F%2Fcafefiles.naver.net%2F20120402_196%2Fsarasa5621_1333335668105taVCH_PNG%2F%25C8%25DE%25C1%25F6%25C5%25EB.png&type=sc960_832")
                await message.channel.send(embed=embed)


client.run("NzY2MDczMjY3NTIxNDU0MTEx.X4eDUw.CrM1KmdQpQ045iCQtWOVEe_TsiA") #여러분들의 토큰값